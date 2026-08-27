"""Структурный разбор файла прайса в строки (specs/content-manager.md §6.7).

В отличие от email_tool.attachments.extract_text (текст с обрезкой 30k для LLM),
здесь возвращаем полные строки таблицы для сигнатуры/маппинга/сопоставления.
Форматы: .xlsx (openpyxl, data_only), .xls (xlrd), .csv. Ошибки не бросаются.
"""
import csv
import io
import logging
from dataclasses import dataclass, field
from pathlib import PurePosixPath

logger = logging.getLogger(__name__)


@dataclass
class Sheet:
    name: str
    rows: list[list[str]] = field(default_factory=list)


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


IMAGE_MARKER = "⟨ИЗОБРАЖЕНИЕ/БАННЕР — возможен разделитель бренда или раздела⟩"


def _from_xlsx(content: bytes) -> list[Sheet]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets: list[Sheet] = []
    try:
        for ws in wb.worksheets:
            rows = [
                [_cell(c) for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            sheets.append(Sheet(name=ws.title, rows=rows))
    finally:
        wb.close()
    return sheets


def image_anchor_rows(content: bytes) -> dict[str, list[int]]:
    """Строки-якоря встроенных изображений по листам (1-based).

    Баннеры брендов часто вставляют картинкой, а не текстом — read_only их не видит,
    поэтому грузим книгу обычным режимом. Возвращает {имя_листа: [номера строк]}.
    """
    from openpyxl import load_workbook

    out: dict[str, list[int]] = {}
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception:
        return out
    try:
        for ws in wb.worksheets:
            rows = []
            for im in getattr(ws, "_images", []):
                try:
                    rows.append(im.anchor._from.row + 1)
                except Exception:
                    continue
            if rows:
                out[ws.title] = sorted(rows)
    finally:
        wb.close()
    return out


def _media_type(data: bytes) -> str:
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if data[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    return "image/png"


def _image_bytes(im) -> bytes | None:
    try:
        data = im._data()  # openpyxl >= 3
        if data:
            return data
    except Exception:
        pass
    ref = getattr(im, "ref", None)
    try:
        if hasattr(ref, "getvalue"):
            return ref.getvalue()
        if isinstance(ref, (str, bytes)):
            with open(ref, "rb") as f:
                return f.read()
    except Exception:
        pass
    return None


def extract_images(content: bytes) -> dict[str, list[tuple[int, bytes, str]]]:
    """Встроенные изображения по листам: {лист: [(строка_1based, байты, media_type)]}.

    Для проверки, что «баннер» — действительно название бренда (агент смотрит на картинку).
    """
    from openpyxl import load_workbook

    out: dict[str, list[tuple[int, bytes, str]]] = {}
    try:
        wb = load_workbook(io.BytesIO(content))
    except Exception:
        return out
    try:
        for ws in wb.worksheets:
            found = []
            for im in getattr(ws, "_images", []):
                try:
                    row = im.anchor._from.row + 1
                except Exception:
                    row = 1
                data = _image_bytes(im)
                if data:
                    found.append((row, data, _media_type(data)))
            if found:
                out[ws.title] = sorted(found, key=lambda t: t[0])
    finally:
        wb.close()
    return out


def mark_images(sheets: list[Sheet], anchors: dict[str, list[int]],
                label: str = IMAGE_MARKER) -> list[Sheet]:
    """Вставляет строки-маркеры на позиции картинок. anchors: {лист: [строки_1based]}."""
    by_name = {s.name: s for s in sheets}
    for name, rows in anchors.items():
        s = by_name.get(name)
        if not s:
            continue
        for r in sorted(rows, reverse=True):   # с конца, чтобы индексы не сдвигались
            idx = min(max(r - 1, 0), len(s.rows))
            s.rows.insert(idx, [label])
    return sheets


def _from_xls(content: bytes) -> list[Sheet]:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    sheets: list[Sheet] = []
    for sh in wb.sheets():
        rows = [
            [_cell(sh.cell_value(i, j)) for j in range(sh.ncols)]
            for i in range(sh.nrows)
        ]
        sheets.append(Sheet(name=sh.name, rows=rows))
    return sheets


def _from_csv(content: bytes) -> list[Sheet]:
    for enc in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode("utf-8", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = [[c.strip() for c in row] for row in csv.reader(io.StringIO(text), dialect)]
    return [Sheet(name="csv", rows=rows)]


def parse_price_table(content: bytes, filename: str) -> list[Sheet]:
    """Разбирает прайс в список листов со строками. Пустой список — если не разобрать."""
    ext = PurePosixPath(filename.lower()).suffix
    try:
        if ext == ".xlsx":
            return _from_xlsx(content)
        if ext == ".xls":
            return _from_xls(content)
        if ext == ".csv":
            return _from_csv(content)
    except Exception:
        logger.exception("Не удалось разобрать прайс %s", filename)
        return []
    logger.warning("Формат %s не поддержан парсером прайса", ext)
    return []


def _rtrim(row: list[str]) -> list[str]:
    """Убирает хвостовые пустые ячейки (листы бывают «широкими» — сотни пустых колонок)."""
    i = len(row)
    while i > 0 and not row[i - 1]:
        i -= 1
    return row[:i]


def non_empty_rows(sheet: Sheet) -> list[list[str]]:
    return [_rtrim(r) for r in sheet.rows if any(c for c in r)]


MAX_PREVIEW_ROWS = 1500


def find_rows(sheet: Sheet, needle: str, limit: int = 60, max_cols: int = 40) -> str:
    """Номера строк листа, где встречается подстрока.

    Для мультибрендового прайса на тысячи строк листать его целиком бессмысленно: каждая
    выгрузка оседает в истории диалога. Дешевле найти, с какой строки начинается раздел
    нужного бренда, и прочитать только его.
    """
    rows = non_empty_rows(sheet)
    want = (needle or "").strip().lower()
    if not want:
        return f"=== Лист: {sheet.name}: пустой запрос поиска ==="
    hits = [(n, r) for n, r in enumerate(rows, 1)
            if any(want in (c or "").lower() for c in r)]
    head = (f"=== Лист: {sheet.name}: «{needle}» найдено в {len(hits)} строках "
            f"из {len(rows)} ===")
    if not hits:
        return head
    lines = [head]
    for n, r in hits[:limit]:
        lines.append(f"{n}\t" + "\t".join(r[:max_cols]))
    if len(hits) > limit:
        lines.append(f"... (показаны первые {limit} совпадений из {len(hits)})")
    lines.append("Чтобы прочитать раздел, вызови read_price_file с from_row = номер "
                 "первой строки раздела.")
    return "\n".join(lines)


def render_preview(sheet: Sheet, max_rows: int = MAX_PREVIEW_ROWS, max_cols: int = 40,
                   start: int = 1) -> str:
    """Таб-текст листа для агента: без хвостовых пустых ячеек, лимит строк и ячеек в строке.

    max_cols защищает от «битых» строк с тысячами дублированных ячеек (ошибки экспорта),
    которые иначе съедают весь бюджет текста.

    `start` (нумерация с 1, по НЕПУСТЫМ строкам) позволяет дочитать длинный лист. Без него
    хвост прайса был недостижим в принципе: прежний лимит в 250 строк резал файл на 394
    строки, и коллекции ниже среза для сопоставления просто не существовали.
    """
    rows = non_empty_rows(sheet)
    start = max(1, start)
    chunk = rows[start - 1:start - 1 + max_rows]
    last = start - 1 + len(chunk)
    shown = (f"{len(rows)} непустых строк" if len(chunk) == len(rows)
             else f"строки {start}\u2013{last} из {len(rows)}")
    lines = [f"=== Лист: {sheet.name} ({shown}) ==="]
    for r in chunk:
        line = "\t".join(r[:max_cols])
        if len(r) > max_cols:
            line += f"\t…(+{len(r) - max_cols} ячеек)"
        lines.append(line)
    if last < len(rows):
        lines.append(f"... (ещё {len(rows) - last} строк НЕ показано. Дочитай их: "
                     f"read_price_file с sheet=«{sheet.name}» и from_row={last + 1})")
    return "\n".join(lines)
